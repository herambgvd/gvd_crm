"""
Import Engine — universal CSV/Excel/Google Sheets import service.

Handles parsing, column mapping, value coercion, and bulk insert
for leads, customers, entities, products, and tickets.
"""

import uuid
import io
import csv
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from core.database import get_database

# ─────────────────── Field Definitions ───────────────────

ENTITY_FIELDS: Dict[str, Dict[str, Dict[str, Any]]] = {
    "lead": {
        "source":              {"type": "string", "required": True},
        "project_name":        {"type": "string"},
        "priority":            {"type": "string", "default": "medium"},
        "expected_value":      {"type": "number"},
        "expected_close_date": {"type": "date"},
        "notes":               {"type": "string"},
        "assigned_to":         {"type": "string"},
    },
    "customer": {
        "company_name":    {"type": "string", "required": True},
        "contact_person":  {"type": "string", "required": True},
        "phone":           {"type": "string"},
        "alternate_phone": {"type": "string"},
        "email":           {"type": "string"},
        "address":         {"type": "string"},
        "city":            {"type": "string"},
        "state":           {"type": "string"},
        "pincode":         {"type": "string"},
        "country":         {"type": "string", "default": "India"},
        "gstin":           {"type": "string"},
        "pan":             {"type": "string"},
        "website":         {"type": "string"},
        "notes":           {"type": "string"},
    },
    "entity": {
        "entity_type":     {"type": "string", "required": True},
        "company_name":    {"type": "string", "required": True},
        "contact_person":  {"type": "string", "required": True},
        "email":           {"type": "string"},
        "phone":           {"type": "string"},
        "alternate_phone": {"type": "string"},
        "address":         {"type": "string"},
        "city":            {"type": "string"},
        "state":           {"type": "string"},
        "pincode":         {"type": "string"},
        "country":         {"type": "string", "default": "India"},
        "gstin":           {"type": "string"},
        "pan":             {"type": "string"},
        "website":         {"type": "string"},
        "notes":           {"type": "string"},
    },
    "product": {
        "name":            {"type": "string", "required": True},
        "sku":             {"type": "string", "required": True},
        "category":        {"type": "string", "required": True},
        "subcategory":     {"type": "string"},
        "description":     {"type": "string"},
        "unit_price":      {"type": "number", "required": True},
        "cost_price":      {"type": "number"},
        "currency":        {"type": "string", "default": "INR"},
        "unit_of_measure": {"type": "string", "default": "piece"},
        "total_quantity":  {"type": "integer", "default": 0},
        "min_stock_level": {"type": "integer", "default": 0},
        "reorder_point":   {"type": "integer", "default": 0},
    },
    "ticket": {
        "customer_name":   {"type": "string", "required": True},
        "title":           {"type": "string"},
        "description":     {"type": "string"},
        "ticket_type":     {"type": "string"},
        "product_name":    {"type": "string"},
        "model_number":    {"type": "string"},
        "customer_email":  {"type": "string"},
        "customer_phone":  {"type": "string"},
        "project_name":    {"type": "string"},
        "location_site":   {"type": "string"},
        "priority":        {"type": "string", "default": "Medium"},
        "category":        {"type": "string", "default": "general"},
    },
}

COLLECTION_MAP = {
    "lead": "leads",
    "customer": "customers",
    "entity": "entities",
    "product": "products",
    "ticket": "tickets",
}


# ─────────────────── Parsing helpers ───────────────────

def parse_csv_content(content: bytes) -> Tuple[List[str], List[List[str]]]:
    """Parse CSV bytes and return (headers, rows)."""
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    headers = [h.strip() for h in rows[0]]
    data_rows = [[c.strip() for c in r] for r in rows[1:] if any(c.strip() for c in r)]
    return headers, data_rows


def parse_excel_content(content: bytes) -> Tuple[List[str], List[List[str]]]:
    """Parse Excel (.xlsx) bytes and return (headers, rows)."""
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([str(c).strip() if c is not None else "" for c in row])
    wb.close()
    if not all_rows:
        return [], []
    headers = all_rows[0]
    data_rows = [r for r in all_rows[1:] if any(c for c in r)]
    return headers, data_rows


# ─────────────────── Value coercion ───────────────────

def coerce_value(value: str, target_type: str) -> Tuple[Any, Optional[str]]:
    """
    Convert a string value to the target type.
    Returns (coerced_value, warning_or_None).
    """
    if not value:
        return None, None

    if target_type == "string":
        return value, None

    if target_type == "number":
        try:
            return float(value.replace(",", "")), None
        except (ValueError, TypeError):
            return None, f"Could not convert '{value}' to number"

    if target_type == "integer":
        try:
            return int(float(value.replace(",", ""))), None
        except (ValueError, TypeError):
            return None, f"Could not convert '{value}' to integer"

    if target_type == "date":
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(value, fmt).isoformat(), None
            except ValueError:
                continue
        return None, f"Could not parse date '{value}'"

    if target_type == "boolean":
        if value.lower() in ("true", "yes", "1", "y"):
            return True, None
        if value.lower() in ("false", "no", "0", "n"):
            return False, None
        return None, f"Could not convert '{value}' to boolean"

    return value, None


# ─────────────────── Ticket number generator ───────────────────

def _generate_ticket_number() -> str:
    return f"TKT-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:8].upper()}"


# ─────────────────── Main import executor ───────────────────

async def execute_import(
    entity_type: str,
    headers: List[str],
    rows: List[List[str]],
    column_mapping: Dict[str, str],
    user_id: str,
) -> Dict[str, Any]:
    """
    Execute bulk import.

    Args:
        entity_type: one of lead, customer, entity, product, ticket
        headers: column headers from the sheet
        rows: data rows (list of lists)
        column_mapping: {sheet_column_name: system_field_name}
        user_id: ID of the importing user

    Returns:
        {imported, total_rows, skipped, warnings}
    """
    if entity_type not in ENTITY_FIELDS:
        raise ValueError(f"Unknown entity type: {entity_type}")

    field_defs = ENTITY_FIELDS[entity_type]
    collection_name = COLLECTION_MAP[entity_type]
    db = get_database()
    collection = db[collection_name]

    # Build header-index lookup
    header_index = {h: i for i, h in enumerate(headers)}

    # For products, pre-fetch existing SKUs for uniqueness check
    existing_skus: set = set()
    if entity_type == "product":
        cursor = collection.find({"is_deleted": {"$ne": True}}, {"sku": 1})
        async for doc in cursor:
            if doc.get("sku"):
                existing_skus.add(doc["sku"])

    now = datetime.now(timezone.utc).isoformat()
    documents: List[Dict[str, Any]] = []
    warnings: List[str] = []
    skipped = 0
    new_skus: set = set()  # track SKUs within this import batch

    for row_idx, row in enumerate(rows, start=2):  # start=2 because row 1 is header
        doc: Dict[str, Any] = {}
        row_warnings: List[str] = []

        # Map columns
        for sheet_col, sys_field in column_mapping.items():
            if sys_field not in field_defs:
                continue  # skip unknown fields silently
            col_index = header_index.get(sheet_col)
            if col_index is None or col_index >= len(row):
                continue
            raw = row[col_index]
            field_def = field_defs[sys_field]
            value, warning = coerce_value(raw, field_def.get("type", "string"))
            if warning:
                row_warnings.append(f"Row {row_idx}, field '{sys_field}': {warning}")
            if value is not None:
                doc[sys_field] = value

        # Apply defaults
        for field_name, field_def in field_defs.items():
            if field_name not in doc and "default" in field_def:
                doc[field_name] = field_def["default"]

        # Check required fields
        missing = [
            f for f, fd in field_defs.items()
            if fd.get("required") and not doc.get(f)
        ]
        if missing:
            skipped += 1
            warnings.append(
                f"Row {row_idx}: skipped — missing required field(s): {', '.join(missing)}"
            )
            continue

        # Product SKU uniqueness
        if entity_type == "product":
            sku = doc.get("sku", "")
            if sku in existing_skus or sku in new_skus:
                skipped += 1
                warnings.append(f"Row {row_idx}: skipped — duplicate SKU '{sku}'")
                continue
            new_skus.add(sku)

        # Ticket number auto-generation
        if entity_type == "ticket":
            doc["ticket_number"] = _generate_ticket_number()
            doc["status"] = "new"

        # Standard metadata
        doc["id"] = str(uuid.uuid4())
        doc["created_by"] = user_id
        doc["created_at"] = now
        doc["updated_at"] = now
        doc["is_deleted"] = False

        warnings.extend(row_warnings)
        documents.append(doc)

    # Bulk insert
    imported = 0
    if documents:
        result = await collection.insert_many(documents)
        imported = len(result.inserted_ids)

    return {
        "imported": imported,
        "total_rows": len(rows),
        "skipped": skipped,
        "warnings": warnings,
    }
